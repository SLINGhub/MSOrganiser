# coding: utf-8

#TODO
#Do it for other machines
#Plotting of graph
#CoV calculation


from MSRawData import AgilentMSRawData
from MSRawData import SciexMSRawData
from MSCalculate import ISTD_Operations
from logging.handlers import TimedRotatingFileHandler
import argparse
import os
import json
from gooey import Gooey
from gooey import GooeyParser
import sys
import logging
import pandas as pd

from datetime import datetime
from pandas import ExcelWriter
from openpyxl import Workbook
from openpyxl.utils import get_column_letter

import warnings
from jinja2 import Environment, FileSystemLoader

import time

def is_frozen():
  return getattr(sys, 'frozen', False)

#We need to change the IO errror
#C:\Users\bchjjs\AppData\Local\Temp
def get_report_dir(dir_name):
    if is_frozen():
        # MEIPASS explanation:
        # https://pythonhosted.org/PyInstaller/#run-time-operation
        basedir = getattr(sys, '_MEIPASS', None)
        if not basedir:
            basedir = os.path.dirname(sys.executable)
        resource_dir = os.path.join(basedir, dir_name)
        if not os.path.isdir(resource_dir):
            raise IOError(
                ("Cannot locate MSreport resources. It seems that the program was frozen, "
                 "but resource files were not copied into directory of the executable "
                 "file. Please copy `msreport` folders from gooey module "
                 "directory into `{}{}` directory. Using PyInstaller, a.datas in .spec "
                 "file must be specified.".format(resource_dir, os.sep)))
        return resource_dir
    else:
        resource_dir = os.path.dirname(__file__)
    return os.path.join(resource_dir, dir_name)


report_dir = get_report_dir('msreport')
env = Environment(loader=FileSystemLoader(report_dir))
ISTD_report_template = env.get_template("ISTD_Report.html")
Parameters_report_template = env.get_template("Parameters_Report.html")
stylesheet_file = os.path.join(report_dir,"typography.css")


dllspath = get_report_dir('cairo_dll')
os.environ['PATH'] = dllspath + os.pathsep + os.environ['PATH']

#To remove the @font-face not available in Windows warning
with warnings.catch_warnings():
    warnings.filterwarnings("ignore", category=UserWarning)
    from weasyprint import HTML

@Gooey(program_name='MS Data Organiser',required_cols=1,optional_cols=1,advanced=True,default_size=(610,710),group_by_type=False)
def parse_args():
    """ Use ArgParser to build up the arguments we will use in our script"""

    """ Save the arguments in a default json file so that we can retrieve them every time we run the script."""

    stored_args = {}
    # get the script name without the extension & use it to build up
    # the json filename
    script_name = os.path.splitext(os.path.basename(__file__))[0]
    args_file =  "{}-args.json".format(script_name)

    # Read in the prior arguments as a dictionary
    if os.path.isfile(args_file):
        with open(args_file) as data_file:
            stored_args = json.load(data_file)

    parser = GooeyParser(description='Create summary tables from MassHunter csv files')

    if not stored_args.get('Output_Format'):
        Output_Format = 'Excel'
    else:
        Output_Format = stored_args.get('Output_Format')

    if not stored_args.get('Transpose_Results'):
        Transpose_Results = 'False'
    else:
        Transpose_Results = stored_args.get('Transpose_Results')

    #Required Arguments 
    parser.add_argument('MS_Files',action='store',nargs="+",help="Input the MS raw files.\nData File is a required column for MassHunter", widget="MultiFileChooser",default=stored_args.get('MS_Files'))
    #parser.add_argument('MS_Files_Type',action='store',nargs="+",help="Input the MS raw files.\nData File is a required column for MassHunter", widget="MultiFileChooser",default=stored_args.get('MS_Files'))
    parser.add_argument('Output_Directory',action='store', help="Output directory to save summary report.", widget="DirChooser",default=stored_args.get('Output_Directory'))
    parser.add_argument('--Output_Format', required=True, choices=['Excel'], help='Select specific file type to output', default=Output_Format)
    parser.add_argument('--Transpose_Results',required=True, choices=['True','False'], help='Set this option to True to let the samples be the columns instead of the Transition_Names',default=Transpose_Results)
    
    #Optional Arguments 
    parser.add_argument('--ISTD_Map', action='store', help='Input the ISTD map file. Required for normalisation', widget="FileChooser",default=stored_args.get('ISTD_Map'))
    #parser.add_argument('--Output_Options', choices=['Area','normArea by ISTD','normConc by ISTD','RT','FWHM','S/N','Precursor Ion','Product Ion'], nargs="+", help='Select specific information to output', widget="Listbox", default=stored_args.get('Output_Options'))
    parser.add_argument('--Output_Options', choices=['Area','normArea by ISTD','RT','FWHM','S/N','Precursor Ion','Product Ion'], nargs="+", help='Select specific information to output', widget="Listbox", default=stored_args.get('Output_Options'))
    parser.add_argument('--Testing', action='store_true', help='Testing mode will generate more output tables.')

    #parser.add_argument('--Transpose_Results', action='store_false', help='Select this option to let the samples be the columns',default=defaults.get('Transpose_Results'))

    #We need this if exe file is called in a UNC path
    if not stored_args.get('Log_Directory'):
        logdirectory = os.path.dirname(os.path.abspath(__file__))
    else:
        logdirectory = stored_args.get('Log_Directory')
    parser.add_argument('--Log_Directory', action='store', help="Directory to save log files.", widget="DirChooser",default=logdirectory)

    #We need this if exe file is called in a UNC path
    if not stored_args.get('Settings_Directory'):
        settings_directory = os.path.dirname(os.path.abspath(__file__))
    else:
        settings_directory = stored_args.get('Settings_Directory')
    parser.add_argument('--Settings_Directory', action='store', help="Directory to save settings", widget="DirChooser",default=settings_directory)

    args = parser.parse_args()

    #Convert the string in Transpose Results to boolean
    if args.Transpose_Results == 'True':
        args.Transpose_Results = True
    else:
        args.Transpose_Results = False

    #Join the list of file paths into one string separated by semicolon
    args_dict = vars(args)
    args_dict['MS_Files'] = ";".join(args_dict['MS_Files'])

    #Check if Output_Options is selected
    if not args_dict['Output_Options']:
        print("Please key in at least one result to output",flush=True)
        sys.exit(-1)

    # Store the values of the arguments so we have them next time we run
    args_file = os.path.join(args_dict['Settings_Directory'],args_file)
    try:
        with open(args_file, 'w') as data_file:
            # Using vars(args) returns the data as a dictionary
            json.dump(vars(args), data_file)
    except Exception as e:
        print("Warning: Unable to save input settings in " + args_file + " due to this error message",flush=True)
        print(e,flush=True)
        #sys.exit(-1)

    return args

def start_logger(log_directory_path):

    logdirectory = os.path.join(log_directory_path,"logfiles")

    try:
        os.makedirs(logdirectory, exist_ok=True)
    except Exception as e:
        print("Unable to create log directory in " + logdirectory + " due to this error message",flush=True)
        print(e,flush=True)
        sys.exit(-1)

    logger = logging.getLogger("MSOrganiser")
    logger.setLevel(logging.INFO)

    # create file handler (fh)
    '''
    logfilename = os.path.join(logdirectory ,datetime.now().strftime('Test_%Y_%m_%d.log'))

    try:
        fh = logging.FileHandler(logfilename)
        fh.setLevel(logging.INFO)
    except Exception as e:
        print("Unable to create log file in " + logfilename + " due to this error message")
        print(e)
        sys.exit(-1)
    '''
    
    logfilename = os.path.join(logdirectory , 'Test_Log')
    #handler = TimedRotatingFileHandler(logfilename,when="h",interval=1,backupCount=0)
    fh = TimedRotatingFileHandler(logfilename,
                                       when='midnight',
                                       interval=1,
                                       backupCount=1)
    
    # create a logging format
    formatter = logging.Formatter('%(asctime)s - %(name)s - %(levelname)s - %(message)s')
    fh.setFormatter(formatter)
    
    # add the handlers to the logger
    logger.addHandler(fh)
    return logger

def start_writer(output_format,output_directory,output_filename):
    """Function to start a writer object"""
    if output_format=="Excel":
        writer = os.path.join(output_directory, output_filename + '_Results.xlsx' )
        #writer = ExcelWriter(writer,engine='xlsxwriter')
        writer = ExcelWriter(writer,engine='openpyxl')
    elif output_format=="csv":
        writer = os.path.join(output_directory, output_filename)
    return writer

def transpose_MSdata(MS_df):
    #Transpose the data
    MS_df = MS_df.T
    #Now the first column is the index, we need to convert back to a first column
    MS_df.reset_index(level=0, inplace=True)
    #Assign the column names from first row
    colnames = MS_df.iloc[0,:].astype('str').str.strip()
    MS_df.columns = colnames
    #We remove the first row because the column names are given
    MS_df = MS_df.iloc[1:]
    #Rename the first column to Compound Name
    MS_df.rename(columns={'Sample_Name':'Transition_Name'}, inplace=True)
    return MS_df

def df_to_file(writer,output_format,output_option,df,logger=None,ingui=False,transpose=False):
    """Funtion to write a df to a file"""

    #Replace '/' as excel cannot have this as sheet title
    if output_option == 'S/N':
        output_option = output_option.replace('/','_to_')

    #If df is empty we send a warning and skip the df
    if df.empty:
        if logger:
            logger.warning('%s has no data. Please check the input file',output_option)
        if ingui:
            print(output_option + ' has no data. Please check the input file',flush=True)
        return

    if transpose:
        df = transpose_MSdata(df)
   
    #Output file into Excel
    if output_format=="Excel":
        try:
            df.to_excel(excel_writer=writer,sheet_name=output_option, index=False)
            worksheet = writer.sheets[output_option]
            for i, width in enumerate(get_col_widths(df)):
                #print(get_column_letter(i+1))
                #print(width)
                #worksheet.set_column(i-1, i-1, width)
                #worksheet.column_dimensions[get_column_letter(i+1)].width = width
                if i==0:
                    continue
                worksheet.column_dimensions[get_column_letter(i)].width = width + 1
        except Exception as e:
            print("Unable to write df to excel file",flush=True)
            print(e,flush=True)
    elif output_format=="csv":
         df.to_csv(writer + output_option + '_Results.csv',sep=',',index=False)

def get_col_widths(dataframe):
    """Function to get the correct width to output the excel file nicely"""
    # First we find the maximum length of the index column   
    idx_max = max([len(str(s)) for s in dataframe.index.values] + [len(str(dataframe.index.name))])
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) + 1 for col in dataframe.columns]

def end_writer(writer,output_format,logger=None,ingui=False):
    """Function to close a writer object"""
    if output_format=="Excel":
        try:
            writer.save()
        except Exception as e:
            if ingui:
                print('Unable to save Excel file due to:',flush=True)
                print(e,flush=True)
            if logger:
                logger.error('Unable to save Excel file due to:')
                logger.error(e)
            sys.exit(-1)

def get_Parameters_df(conf,MS_File):
    Parameter_list = []
    args_dict = vars(conf)

    #Get specific keys into the parameters list
    Parameter_list.append(("Input_File",os.path.basename(MS_File)))

    if args_dict['Output_Format']:
        Parameter_list.append(("Output_Format",args_dict['Output_Format']))

    if args_dict['ISTD_Map']:
        Parameter_list.append(("ISTD_Map_File",os.path.basename(args_dict['ISTD_Map'])))

    if args_dict['Output_Options']:
        for things in args_dict['Output_Options']:
            Parameter_list.append(("Output_Options",things))

    Parameters_df = pd.DataFrame(Parameter_list,columns=['Parameters', 'Value'])
    return Parameters_df

def create_timed_rotating_log(path):
    """"""
    logger = logging.getLogger("Rotating Log")
    logger.setLevel(logging.INFO)
    
    handler = TimedRotatingFileHandler(path,
                                       when="m",
                                       interval=1,
                                       backupCount=5)
    logger.addHandler(handler)
    for i in range(6):
        logger.info("This is a test!")
        time.sleep(75)

if __name__ == '__main__':

    #Read the parser
    conf = parse_args()

    #Start log on a job
    logger = start_logger(conf.Log_Directory)
    logger.info("Starting the job.")
    print("Starting the job.",flush=True)

    #log_file = "timed_test.log"
    #create_timed_rotating_log(log_file)

    #We do this for every mass hunter file output
    for MS_File in conf.MS_Files.split(';'):

        print("Working on " + MS_File,flush=True)
        logger.info("Working on " + MS_File)

        #Initiate the ISTD report
        pdfs = []

        #Generate the parameters report
        Parameters_df = get_Parameters_df(conf,MS_File)
        template_vars = {"title": "Parameters", "Parameter_Report": Parameters_df.to_html(index=False)}
        html_string = Parameters_report_template.render(template_vars)
        html = HTML(string=html_string)
        pdfs.append(html.render(stylesheets=[stylesheet_file]))

        #Check if file is from Agilent or Sciex
        if MS_File.endswith('.csv'):
            RawData = AgilentMSRawData(filepath=MS_File,logger=logger)
        elif MS_File.endswith('.txt'):
            RawData = SciexMSRawData(filepath=MS_File,logger=logger)
        output_filename = os.path.splitext(os.path.basename(MS_File))[0]

        #Set up the file writing configuration for Excel, ...
        writer = start_writer(conf.Output_Format,conf.Output_Directory,output_filename)
        
        if conf.Output_Options:
            for column_name in conf.Output_Options:
                if column_name == 'normArea by ISTD':
                    #Perform normalisation using ISTD
                    #Get Area Table
                    Area_df = RawData.get_table('Area',is_numeric=True)
                    #Get ISTD map df
                    ISTD_map_df = ISTD_Operations.read_ISTD_map(conf.ISTD_Map,column_name,logger,ingui=True)
                    #Perform normalisation using ISTD
                    [norm_Area_df,ISTD_Area,ISTD_Report] = ISTD_Operations.normalise_by_ISTD(Area_df,ISTD_map_df,logger,ingui=True)
                    #Output the normalised area results
                    if conf.Testing:
                        df_to_file(writer,conf.Output_Format,"ISTD Area",ISTD_Area,logger,ingui=True,transpose=conf.Transpose_Results)
                    df_to_file(writer,conf.Output_Format,"ISTD map",ISTD_map_df,logger,ingui=True)
                    df_to_file(writer,conf.Output_Format,column_name,norm_Area_df,logger,ingui=True,transpose=conf.Transpose_Results)

                    #Print ISTD report to pdf when it is not empty
                    if not ISTD_Report.empty:
                        template_vars = {"title": "ISTD_Report","column_name": column_name, "ISTD_Report": ISTD_Report.to_html()}
                        html_string = ISTD_report_template.render(template_vars)
                        html = HTML(string=html_string)
                        pdfs.append(html.render(stylesheets=[stylesheet_file]))
                elif column_name == 'normConc by ISTD':
                    try:
                        norm_Area_df
                    except NameError:
                        #Perform normalisation using ISTD
                        #Get Area Table
                        Area_df = RawData.get_table('Area',is_numeric=True)
                        #Get ISTD map df
                        ISTD_map_df = ISTD_Operations.read_ISTD_map(conf.ISTD_Map,column_name,logger,ingui=True)
                        #Perform normalisation using ISTD
                        [norm_Area_df,ISTD_Area,ISTD_Report] = ISTD_Operations.normalise_by_ISTD(Area_df,ISTD_map_df,logger,ingui=True)

                        #Print ISTD report to pdf when it is not empty
                        if not ISTD_Report.empty:
                            template_vars = {"title": "ISTD_Report","column_name": column_name, "ISTD_Report": ISTD_Report.to_html()}
                            html_string = ISTD_report_template.render(template_vars)
                            html = HTML(string=html_string)
                            pdfs.append(html.render(stylesheets=[stylesheet_file]))

                    #Perform concentration calculation
                    [norm_Conc_df,ISTD_Conc_df,ISTD_Samp_Ratio_df] = ISTD_Operations.getConc_by_ISTD(norm_Area_df,ISTD_map_df,logger,ingui=True)
                    if conf.Testing:
                        df_to_file(writer,conf.Output_Format,"ISTD Conc",ISTD_Conc_df,logger,ingui=True,transpose=conf.Transpose_Results)
                        df_to_file(writer,conf.Output_Format,"ISTD to Samp Vol Ratio",ISTD_Samp_Ratio_df,logger,ingui=True,transpose=conf.Transpose_Results)
                    df_to_file(writer,conf.Output_Format,"ISTD map",ISTD_map_df,logger,ingui=True)
                    df_to_file(writer,conf.Output_Format,column_name,norm_Conc_df,logger,ingui=True,transpose=conf.Transpose_Results)

                else:
                    #We extract the data directly from the file and output accordingly
                    Output_df = RawData.get_table(column_name,is_numeric=True)
                    df_to_file(writer,conf.Output_Format,column_name,Output_df,logger,ingui=True,transpose=conf.Transpose_Results)

        #End the writing configuration for Excel, ...
        end_writer(writer,conf.Output_Format,logger,ingui=True)

        #Output the ISTD report for each file
        val = []
        for doc in pdfs:
            for p in doc.pages:
                val.append(p)
        #print(pdfs[0].copy(val))
        pdf_file = pdfs[0].copy(val).write_pdf(os.path.join(conf.Output_Directory , output_filename + "_Report.pdf")) # use metadata of first pdf

    #End log on a job
    logger.info("Job is finished.")
    print("Job is finished",flush=True)


#Jeremy's style
"""
#Separate the Sample data
Sampledf = df.loc[:, df.iloc[0,:].str.contains("Sample")]
#Remove the whitespace in the column names at row 2 and assign it to the Sampledf column name
colnames = Sampledf.iloc[1,:].astype('str').str.strip()
Sampledf.columns = colnames
#We remove the first and second row because the column names are given
Sampledf = Sampledf.iloc[2:]
#Reset the row index
Sampledf = Sampledf.reset_index(drop=True)
#Strip the whitespaces for each string columns
Sampledf = remove_whiteSpaces(Sampledf)

#Separate the Compound Results data
CompResultsdf = df.loc[:, df.iloc[0,:].str.contains("Results")]
#Remove the text "Results" and whitespaces
CompResultsdf.iloc[0,:] = CompResultsdf.iloc[0,:].str.replace("Results", "").str.strip()
CompResultsdf.iloc[1,:] = CompResultsdf.iloc[1,:].str.strip()
#Rearrange the columns names based on the order 
# Comp1.Result1, Comp2.Result1, Comp1.Result2, Comp2.Result2, . . .
sorted_colname = sorted(zip(CompResultsdf.iloc[0,:].astype('str'), CompResultsdf.iloc[1,:].astype('str')) , key=lambda name:name[1])
sorted_colname = ['.'.join(w) for w in sorted_colname]
#Assign column name to CompResultsdf
colnames = CompResultsdf.iloc[0,:].astype('str') + "." + CompResultsdf.iloc[1,:].astype('str')
CompResultsdf.columns = colnames
#We remove the first and second row because the column names are given
CompResultsdf = CompResultsdf.iloc[2:]
#Reset the row index
CompResultsdf = CompResultsdf.reset_index(drop=True)
#Reorder the columns of CompResultsdf to the order of sorted_colname
CompResultsdf = CompResultsdf.ix[:, sorted_colname]
#Strip the whitespaces for each string columns
CompResultsdf = remove_whiteSpaces(CompResultsdf)


#Separate the Compound Method data
CompMethoddf = df.loc[:, df.iloc[0,:].str.contains("Method")]
#Remove the text "Method" and whitespaces
CompMethoddf.iloc[0,:] = CompMethoddf.iloc[0,:].str.replace("Method", "").str.strip()
CompMethoddf.iloc[1,:] = CompMethoddf.iloc[1,:].str.strip()
#Rearrange the columns names based on the order 
# Comp1.Method1, Comp2.Method1, Comp1.Method2, Comp2.Method2, . . .
sorted_colname = sorted(zip(CompMethoddf.iloc[0,:].astype('str'), CompMethoddf.iloc[1,:].astype('str')) , key=lambda name:name[1])
sorted_colname = ['.'.join(w) for w in sorted_colname]
#Assign column name to CompMethoddf
colnames = CompMethoddf.iloc[0,:].astype('str') + "." + CompMethoddf.iloc[1,:].astype('str')
CompMethoddf.columns = colnames
#We remove the first and second row because the column names are given
CompMethoddf = CompMethoddf.iloc[2:]
#Reset the row index
CompMethoddf = CompMethoddf.reset_index(drop=True)
#Reorder the columns of CompResultsdf to the order of sorted_colname
CompMethoddf = CompMethoddf.ix[:,sorted_colname]
#Strip the whitespaces for each string columns
CompMethoddf = remove_whiteSpaces(CompMethoddf)

result = pd.concat([Sampledf, CompMethoddf , CompResultsdf ], axis=1)
"""