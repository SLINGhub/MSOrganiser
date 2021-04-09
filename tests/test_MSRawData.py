import unittest
import os
import pandas as pd
import openpyxl
from MSRawData import AgilentMSRawData
from MSRawData import SciexMSRawData
from MSDataOutput import MSDataOutput

WIDETABLEFORM_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm.csv')
WIDETABLEFORM_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm_Results.xlsx')
WIDETABLEFORM_TRANSPOSE_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm_TransposeResults.xlsx')

WIDETABLEFORM_ISO_8859_1_ENCODE_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'ISO_8859_1_Encoding.csv')
WIDETABLEFORM_ISO_8859_1_ENCODE_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'ISO_8859_1_Encoding_Results.xlsx')

LARGE_WIDETABLEFORM_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'LargeTestData.csv')
LARGE_WIDETABLEFORM_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'LargeTestData_Results.xlsx')

COMPOUNDTABLEFORM_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'CompoundTableForm.csv')
COMPOUNDTABLEFORM_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'CompoundTableForm_Results.xlsx')

SCIEX_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'SciExTestData.txt')
SCIEX_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'SciExTestData_Results.xlsx')

class Agilent_Test(unittest.TestCase):

    def setUp(self):
        self.WideData = AgilentMSRawData(WIDETABLEFORM_FILENAME,ingui=True)
        self.ISO_EncodedData = AgilentMSRawData(WIDETABLEFORM_ISO_8859_1_ENCODE_FILENAME,ingui=True)
        self.LargeWideData = AgilentMSRawData(LARGE_WIDETABLEFORM_FILENAME,ingui=True)
        self.CompoundData = AgilentMSRawData(COMPOUNDTABLEFORM_FILENAME,ingui=True)
        self.SciexData = SciexMSRawData(SCIEX_FILENAME,ingui=True)

        self.WideDataResults = openpyxl.load_workbook(WIDETABLEFORM_RESULTS_FILENAME)
        self.WideDataTransposeResults = openpyxl.load_workbook(WIDETABLEFORM_TRANSPOSE_RESULTS_FILENAME)
        self.ISO_EncodedDataResults = openpyxl.load_workbook(WIDETABLEFORM_ISO_8859_1_ENCODE_RESULTS_FILENAME)
        self.LargeWideDataResults = openpyxl.load_workbook(LARGE_WIDETABLEFORM_RESULTS_FILENAME)
        self.CompoundDataResults = openpyxl.load_workbook(COMPOUNDTABLEFORM_RESULTS_FILENAME)
        self.SciexDataResults = openpyxl.load_workbook(SCIEX_RESULTS_FILENAME)

    def test_WideData(self):
        """Check if the software is able to do the following with WideTableForm.csv:

        * Extract Area, RT and FWHM successfully using AgilentMSRawData.get_table
        * Traspose the data correctly using MSDataOutput.transpose_MSdata
        """

        self.assertEqual("WideTableForm",self.WideData.DataForm)
        self.__compare_tables("Area",self.WideData,self.WideDataResults)
        self.__compare_tables("Area",self.WideData,self.WideDataTransposeResults,transpose=True)
        self.__compare_tables("RT",self.WideData,self.WideDataResults)
        self.__compare_tables("RT",self.WideData,self.WideDataTransposeResults,transpose=True)
        self.__compare_tables("FWHM",self.WideData,self.WideDataResults)
        self.__compare_tables("FWHM",self.WideData,self.WideDataTransposeResults,transpose=True)

    def test_ISOEncodedData(self):
        """Check if the software is able to do the following with ISO_8859_1_Encoding.csv:

        * Extract Area successfully using AgilentMSRawData.get_table
        """

        self.assertEqual("WideTableForm",self.WideData.DataForm)
        self.__compare_tables("Area",self.ISO_EncodedData,self.ISO_EncodedDataResults)


    def test_WideDataLarge(self):
        """Check if the software is able to do the following with large dataset sPerfect_Index_AllLipids_raw.csv:

        * Extract Area, RT and FWHM successfully using AgilentMSRawData.get_table
        """

        self.assertEqual("WideTableForm",self.LargeWideData.DataForm)
        self.__compare_tables("Area",self.LargeWideData,self.LargeWideDataResults)
        self.__compare_tables("RT",self.LargeWideData,self.LargeWideDataResults)
        self.__compare_tables("FWHM",self.LargeWideData,self.LargeWideDataResults)

    def test_CompoundData(self):
        """Check if the software is able to do the following with large dataset CompoundTableForm.csv:

        * Extract Area and RT successfully using AgilentMSRawData.get_table
        """

        self.assertEqual("CompoundTableForm",self.CompoundData.DataForm)
        self.__compare_tables("Area",self.CompoundData,self.CompoundDataResults)
        self.__compare_tables("RT",self.CompoundData,self.CompoundDataResults)

    def test_SciexData(self):
        """Check if the software is able to do the following with large dataset Mohammed_SciEx_data.txt:

        * Extract Area and RT successfully using SciexMSRawData.get_table
        """

        self.__compare_tables("Area",self.SciexData,self.SciexDataResults)
        self.__compare_tables("RT",self.SciexData,self.SciexDataResults)
        self.__compare_tables("FWHM",self.SciexData,self.SciexDataResults)

    def tearDown(self):
        self.WideDataResults.close()
        self.WideDataTransposeResults.close()
        self.LargeWideDataResults.close()
        self.CompoundDataResults.close()
        self.SciexDataResults.close()

    def __compare_tables(self,table_name,MSDataObject,ExcelWorkbook,transpose=False):
        '''Check if the pandas data frame (MSDataObject) has the same values as the table in ExcelWorkbook'''
        ExcelWorkbook = self.__sheet_to_table(ExcelWorkbook,table_name).apply(pd.to_numeric, errors='ignore', downcast = 'float')
        if transpose:
            MSDataObject = MSDataOutput.transpose_MSdata(MSDataObject.get_table(table_name)).apply(pd.to_numeric, errors='ignore', downcast = 'float')
        else:
            MSDataObject = MSDataObject.get_table(table_name).apply(pd.to_numeric, errors='ignore', downcast = 'float')
        pd.testing.assert_frame_equal(MSDataObject,ExcelWorkbook)

    def __sheet_to_table(self,workbook,sheet_name):
        '''Convert an Excel sheet table into a pandas data frame.'''
        ws = workbook[sheet_name]
        data = ws.values
        cols = next(data)
        data = list(data)
        return pd.DataFrame(data, columns=cols)

if __name__ == '__main__':
    unittest.main()
