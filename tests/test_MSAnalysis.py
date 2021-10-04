import unittest
from unittest.mock import patch
import os
import pandas as pd
import openpyxl
from MSAnalysis import MS_Analysis
from MSOrganiser import concatenate_along_rows_workflow

WIDETABLEFORM_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm.csv')
WIDETABLEFORM_ANNOTATION = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm_Annotation.xlsx')
WIDETABLEFORM_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm_Results.xlsx')
WIDETABLEFORM_LONGTABLE_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm_LongTable.xlsx')
WIDETABLEFORM_LONGTABLE_WITH_ANNOT_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'WideTableForm_LongTable_with_Annot.xlsx')

LARGE_WIDETABLEFORM_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'LargeTestData.csv')
LARGE_WIDETABLEFORM_ANNOTATION = os.path.join(os.path.dirname(__file__),"testdata", 'LargeTestData_Annotation.xlsx')
LARGE_WIDETABLEFORM_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'LargeTestData_Results.xlsx')

COMPOUNDTABLEFORM_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'CompoundTableForm.csv')
COMPOUNDTABLEFORM_ANNOTATION = os.path.join(os.path.dirname(__file__),"testdata", 'CompoundTableForm_Annotation.xlsx')
COMPOUNDTABLEFORM_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'CompoundTableForm_Results.xlsx')

SCIEX_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'SciExTestData.txt')
SCIEX_ANNOTATION = os.path.join(os.path.dirname(__file__),"testdata", 'SciExTestData_Annotation.xlsx')
SCIEX_RESULTS_FILENAME = os.path.join(os.path.dirname(__file__),"testdata", 'SciExTestData_Results.xlsx')

class Agilent_Test(unittest.TestCase):

    # See https://realpython.com/lessons/mocking-print-unit-tests/
    # for more details on mock
    def setUp(self):
        # Replace the print function in Annotation.py file to a mock
        self.patcher = patch('MSCalculate.print')

        self.InputDataList = [WIDETABLEFORM_FILENAME,
                              LARGE_WIDETABLEFORM_FILENAME,
                              COMPOUNDTABLEFORM_FILENAME,
                              SCIEX_FILENAME]

        self.DataResultList = [WIDETABLEFORM_RESULTS_FILENAME,
                               LARGE_WIDETABLEFORM_RESULTS_FILENAME,
                               COMPOUNDTABLEFORM_RESULTS_FILENAME,
                               SCIEX_RESULTS_FILENAME]

        self.AnnotationList = [WIDETABLEFORM_ANNOTATION,
                               LARGE_WIDETABLEFORM_ANNOTATION,
                               COMPOUNDTABLEFORM_ANNOTATION,
                               SCIEX_ANNOTATION]

        self.MSFileTypeList = ['Agilent Wide Table in csv',
                               'Agilent Wide Table in csv',
                               'Agilent Compound Table in csv',
                               'Multiquant Long Table in txt']

    def test_getnormAreaTable(self):
        """Check if the software is able to calculate the normalise area using MS_Analysis.get_Normalised_Area from these datasets 

        * WideTableForm.csv
        * LargeTestData.csv
        * CompoundTableForm.csv
        * SciExTestData.txt
        """

        #Perform normalisation using ISTD
        for i in range(len(self.InputDataList)):
            MyData = MS_Analysis(MS_FilePath = self.InputDataList[i],
                                 MS_FileType = self.MSFileTypeList[i],
                                 Annotation_FilePath = self.AnnotationList[i],
                                 ingui = True)
            Results = openpyxl.load_workbook(self.DataResultList[i])

            if os.path.basename(self.InputDataList[i]) == "LargeTestData.csv" :
                mock_print = self.patcher.start()

            [norm_Area_df,ISTD_Area,Transition_Name_Annot,ISTD_Report] = MyData.get_Normalised_Area('normArea by ISTD')
            
            if os.path.basename(self.InputDataList[i]) == "LargeTestData.csv" :
                mock_print.assert_called_with('There are Transition_Names mentioned in the ' +
                                              'Transition_Name_Annot sheet but have a blank Transition_Name_ISTD.\n' +
                                              '\"Sph d17:0 (ISTD)\"\n' + 
                                              '\"Sph d17:1 (ISTD)\"', 
                                              flush = True)

            self.__compare_df('normArea_by_ISTD',norm_Area_df,Results)
            self.__compare_df('Transition_Name_Annot',Transition_Name_Annot,Results)
            Results.close()


    def test_getAnalyteConcTable(self):
        """Check if the software is able to calculate the transition names concentration using MS_Analysis.get_Analyte_Concentration from these datasets

        * WideTableForm.csv
        * LargeTestData.csv
        * CompoundTableForm.csv
        * SciExTestData.txt
        """

        #Perform analyte concentration using ISTD
        for i in range(len(self.InputDataList)):
            MyData = MS_Analysis(MS_FilePath = self.InputDataList[i],
                                 MS_FileType = self.MSFileTypeList[i],
                                 Annotation_FilePath = self.AnnotationList[i],
                                 ingui = True)
            Results = openpyxl.load_workbook(self.DataResultList[i])

            if os.path.basename(self.InputDataList[i]) == "LargeTestData.csv" :
                mock_print = self.patcher.start()

            [norm_Conc_df,ISTD_Conc_df,ISTD_Samp_Ratio_df,Sample_Annot_df] = MyData.get_Analyte_Concentration('normConc by ISTD')
            
            if os.path.basename(self.InputDataList[i]) == "LargeTestData.csv" :
                mock_print.assert_called_with('There are Transition_Names mentioned in the ' +
                                              'Transition_Name_Annot sheet but have a blank Transition_Name_ISTD.\n' +
                                              '\"Sph d17:0 (ISTD)\"\n' + 
                                              '\"Sph d17:1 (ISTD)\"', 
                                              flush = True)
            
            self.__compare_df('normConc_by_ISTD',norm_Conc_df,Results)
            
            #Remove the column "Merge_Status" as it is not relevant
            #Reorder the column such that "Concentration_Unit" is at the last column
            Sample_Annot_df = Sample_Annot_df[["Data_File_Name", "Sample_Name",
                                               "Sample_Amount", "Sample_Amount_Unit",
                                               "ISTD_Mixture_Volume_[uL]", "ISTD_to_Sample_Amount_Ratio",
                                               "Concentration_Unit"]]
            self.__compare_df('Sample_Annot',Sample_Annot_df,Results)
            Results.close()

    def test_getLongTable(self):
        """Check if the software is able to get the LongTable from these datasets

        * WideTableForm.csv
        """

        MyLongTableData = MS_Analysis(MS_FilePath = WIDETABLEFORM_FILENAME,
                                      MS_FileType = 'Agilent Wide Table in csv',
                                      Annotation_FilePath = WIDETABLEFORM_ANNOTATION,
                                      ingui = True,
                                      longtable = True, 
                                      longtable_annot = False)

        LongTableDataResults = openpyxl.load_workbook(WIDETABLEFORM_LONGTABLE_FILENAME)

        MyLongTableData.get_from_Input_Data('Area', outputdata = False)
        Long_Table_df = MyLongTableData.get_Long_Table()
        self.__compare_df('Long_Table',Long_Table_df,LongTableDataResults)

    def test_getLongTable_with_Annot(self):
        """Check if the software is able to get the LongTable with Annotations from these datasets

        * WideTableForm.csv
        """

        MyLongTableDataWithAnnot = MS_Analysis(MS_FilePath = WIDETABLEFORM_FILENAME,
                                               MS_FileType = 'Agilent Wide Table in csv',
                                               Annotation_FilePath = WIDETABLEFORM_ANNOTATION,
                                               ingui = True,
                                               longtable = True, 
                                               longtable_annot = True)
        LongTableDataWithAnnotResults = openpyxl.load_workbook(WIDETABLEFORM_LONGTABLE_WITH_ANNOT_FILENAME)

        MyLongTableDataWithAnnot.get_from_Input_Data('Area', outputdata = False)
        MyLongTableDataWithAnnot.get_Normalised_Area('normArea by ISTD', outputdata = False)
        MyLongTableDataWithAnnot.get_Analyte_Concentration('normConc by ISTD', outputdata = False)
        Long_Table_df = MyLongTableDataWithAnnot.get_Long_Table()
        self.__compare_df('Long_Table',Long_Table_df,LongTableDataWithAnnotResults)
           
    def tearDown(self):
        self.patcher.stop()

    def __compare_df(self,table_name,MSData_df,ExcelWorkbook):
        MSData_df = MSData_df.apply(pd.to_numeric, errors='ignore', downcast = 'float')
        ExcelWorkbook = self.__sheet_to_table(ExcelWorkbook,table_name).apply(pd.to_numeric, errors='ignore', downcast = 'float')
        pd.testing.assert_frame_equal(MSData_df,ExcelWorkbook)

    def __sheet_to_table(self,workbook,sheet_name):
        ws = workbook[sheet_name]
        data = ws.values
        cols = next(data)
        data = list(data)
        return pd.DataFrame(data, columns=cols)

if __name__ == '__main__':
    unittest.main()
