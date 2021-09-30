import pandas as pd
import numpy as np
import sys
import re
import os
import logging

from openpyxl import load_workbook

class MS_Template():
    """A class to describe the excel macro sheet MS Template Creator

    Args:
        filePath (str): file path of the input MS Template Creator file
        logger (object): logger object created by start_logger in MSOrganiser
        ingui (bool): if True, print analysis status to screen
        doing_normalization (bool): if True, check if input file has data. If no data, throws an error
        allow_multiple_istd (bool): if True, allow normalization of data by mulitple internal standards
    """

    def __init__(self,filepath,column_name, logger=None,ingui=True, 
                 doing_normalization = False, allow_multiple_istd = False):
        self.__logger = logger
        self.__ingui = ingui
        self.filepath = filepath
        self.__filecheck(column_name)
        self.__doing_normalization = doing_normalization
        self.__allow_multiple_istd = allow_multiple_istd

    def remove_whiteSpaces(df):
        """Strip the whitespaces for each string columns of a df

        Args:
            df (pandas DataFrame): A panda data frame

        Returns:
            df (pandas DataFrame): A panda data frame with white space removed

        """

        df[df.select_dtypes(['object']).columns] = df.select_dtypes(['object']).apply(lambda x: x.str.strip())
        return df

    def __filecheck(self,column_name):
        # Check if input is blank/None
        if not self.filepath:
            if self.__logger:
                self.__logger.error('A ISTD map file is required to perform this calculation: %s', column_name)
            if self.__ingui:
                print('A ISTD map file is required to perform this calculation: ' + column_name,flush=True)
            sys.exit(-1)

        # Check if file exists
        if not os.path.isfile(self.filepath):
            if self.__logger:
                self.__logger.error('%s does not exists. Please check the input file',self.filepath)
            if self.__ingui:
                print(self.filepath + ' does not exists. Please check the input file',flush=True)
            sys.exit(-1)

        if self.filepath.endswith('.csv'):
            if self.__logger:
                self.__logger.error('This program no longer accept csv file as input for the ISTD map file. Please use the excel template file given')
            if self.__ingui:
                print('This program no longer accept csv file as input for the ISTD map file. Please use the excel template file given',flush=True)
            sys.exit(-1)

    def __readExcelWorkbook(self):
        # Read the excel file
        try:
            wb = load_workbook(filename=self.filepath,data_only=True)
        except Exception as e:
            if self.__logger:
                self.__logger.error("Unable to read excel file %s",self.filepath)
                self.__logger.error(e)
            if self.__ingui:
                print("Unable to read excel file " + self.filepath,flush=True)
                print(e,flush=True)
            sys.exit(-1)
        return wb

    def __checkExcelWorksheet_in_Workbook(self,sheetname,wb):
        # Check if the excel file has the sheet sheetname
        if sheetname not in wb.sheetnames:
            if self.__logger:
                self.__logger.error('Sheet name ' + sheetname + ' does not exists. Please check the input excel file.')
            if self.__ingui:
                print('Sheet name ' + sheetname + ' does not exists. Please check the input excel file.',flush=True)
            sys.exit(-1)

    def __check_if_df_is_empty(self,sheetname,df):
        # Validate the input sheet has data
        if df.empty:
            if self.__logger:
                self.__logger.warning('The input ' + sheetname + ' sheet has no data.')
            if self.__ingui:
                print('The input ' + sheetname + ' sheet has no data.',flush=True)
            sys.exit(-1)

    def __checkColumns_in_df(self,colname,sheetname,df):
        # Check if the column name exists as a header in the df
        if colname not in df:
            if self.__logger:
                self.__logger.error('The ' + sheetname  + ' sheet is missing the column ' + colname + '.')
            if self.__ingui:
                print('The ' + sheetname  + ' sheet is missing the column ' + colname + '.',flush=True)
            sys.exit(-1)

    def __checkDuplicates_in_cols(self,colname_list,sheetname,df):
        # Check if a list of columns in the input df has duplicate data
        duplicateValues = df.duplicated(subset=colname_list)
        if duplicateValues.any():
            duplicatelist = [ str(int(i) + 2)  for i in duplicateValues[duplicateValues==True].index.tolist()]
            if self.__logger:
                self.__logger.error('Data at ' + ', '.join(colname_list) + ' column(s) in the ' + sheetname + 
                                    ' sheet have duplicates at row %s', ', '.join(duplicatelist) + '.')
            if self.__ingui:
                print('Data at ' + ', '.join(colname_list) + ' column(s) in the ' + sheetname + 
                      ' sheet has duplicates at row ' + ', '.join(duplicatelist) + '.',
                      flush=True)
            sys.exit(-1)

    def Read_Transition_Name_Annot_Sheet(self):
        """Read the excel sheet Transition_Name_Annot as a pandas data frame

        Returns:
            Transition_Name_Annot_df (pandas DataFrame): A panda data frame containing the contents of Transition_Name_Annot
        """

        #Open the excel file
        wb = self.__readExcelWorkbook()

        #Check if the excel file has the sheet "Transition_Name_Annot"
        self.__checkExcelWorksheet_in_Workbook("Transition_Name_Annot",wb)
        
        #Convert worksheet to a dataframe
        worksheet = wb["Transition_Name_Annot"]
        #Get the column names in the first row of the excel sheet
        cols = next(worksheet.values)[0:]
        Transition_Name_Annot_df = pd.DataFrame(worksheet.values, columns=cols)

        #We remove the first row as the headers as been set up
        Transition_Name_Annot_df = Transition_Name_Annot_df.iloc[1:]
        #Reset the row index
        Transition_Name_Annot_df = Transition_Name_Annot_df.reset_index(drop=True)

        #Remove rows with all None, NA,NaN
        Transition_Name_Annot_df = Transition_Name_Annot_df.dropna(axis=0, how='all')

        #Validate the Transition_Name_Annot sheet is valid (Has the Transition_Name and Transition_Name_ISTD columns are not empty)
        self.__validate_Transition_Name_Annot_sheet("Transition_Name_Annot",Transition_Name_Annot_df, 
                                                    allow_multiple_istd = self.__allow_multiple_istd)

        #Remove whitespaces in column names
        Transition_Name_Annot_df.columns = Transition_Name_Annot_df.columns.str.strip()

        #Remove whitespace for each string column
        Transition_Name_Annot_df = MS_Template.remove_whiteSpaces(Transition_Name_Annot_df)

        #Remove columns with all None, NA,NaN
        Transition_Name_Annot_df = Transition_Name_Annot_df.dropna(axis=1, how='all')
          

        #print(Transition_Name_Annot_df)

        #Close the workbook
        wb.close()

        return Transition_Name_Annot_df

    def __validate_Transition_Name_Annot_sheet(self,sheetname,Transition_Name_Annot_df,
                                               allow_multiple_istd = False):
        #Validate the Transition_Name_Annot sheet has data when normalization is performed
        if self.__doing_normalization:
            self.__check_if_df_is_empty(sheetname,Transition_Name_Annot_df)

        #Check if the column Transition_Name exists as a header in Transition_Name_Annot_df
        self.__checkColumns_in_df('Transition_Name',sheetname,Transition_Name_Annot_df)

        #Check if Transition_Name column has duplicate Transition_Names
        if allow_multiple_istd:
            self.__checkDuplicates_in_cols(colname_list = ['Transition_Name', 'Transition_Name_ISTD'],
                                           sheetname = sheetname,
                                           df = Transition_Name_Annot_df)
        else:
            self.__checkDuplicates_in_cols(colname_list = ['Transition_Name'],
                                           sheetname = sheetname,
                                           df = Transition_Name_Annot_df)

        #Check if the column Transition_Name exists as a header in Transition_Name_Annot_df
        self.__checkColumns_in_df('Transition_Name_ISTD',sheetname,Transition_Name_Annot_df)

    def Read_ISTD_Annot_Sheet(self):
        """Read the excel sheet ISTD_Annot as a pandas data frame

        Returns:
            ISTD_Annot_df (pandas DataFrame): A panda data frame containing the contents of ISTD_Annot
        Note:
            Transition_Name_ISTD, ISTD_Conc_[nM] and Custom Unit are the only columns taken
        """

        #Open the excel file
        wb = self.__readExcelWorkbook()

        #Check if the excel file has the sheet "ISTD_Annot"
        self.__checkExcelWorksheet_in_Workbook("ISTD_Annot",wb)
        
        #Convert worksheet to a dataframe
        worksheet = wb["ISTD_Annot"]

        #Check that sheet is valid
        self.__validate_ISTD_Annot_Sheet(worksheet)

        #Get the column names
        istd_conc_name = re.sub("\[.*?\]",worksheet["F3"].value,worksheet["E3"].value)
        cols = [worksheet["A2"].value, istd_conc_name]

        #Get the ISTD Table and clean it up
        ISTD_Annot_df = worksheet.values
        ISTD_Annot_df = pd.DataFrame(ISTD_Annot_df)
        
        #We remove the first three row as the headers as been set up
        ISTD_Annot_df = ISTD_Annot_df.iloc[3:]
        #Reset the row index
        ISTD_Annot_df = ISTD_Annot_df.reset_index(drop=True)
        
        #Take specific columns (A and F only)
        ISTD_Annot_df = ISTD_Annot_df.iloc[:,[0,5]]
        ISTD_Annot_df.columns = cols

        #Remove rows with no Transition_Name_ISTD
        ISTD_Annot_df = ISTD_Annot_df.dropna(subset=['Transition_Name_ISTD'])
        #ISTD_Annot_df = ISTD_Annot_df.dropna(axis=0, how='all')

        #Check if Transition_Name_ISTD column has duplicate Transition_Name_ISTD
        self.__checkDuplicates_in_cols(colname_list = ['Transition_Name_ISTD'],
                                       sheetname = 'ISTD_Annot',
                                       df = ISTD_Annot_df)

        #Remove whitespaces in column names
        ISTD_Annot_df.columns = ISTD_Annot_df.columns.str.strip()

        #Convert all but first column to numeric
        #ISTD_Annot_df['ISTD_Conc_[nM]'] = pd.to_numeric(ISTD_Annot_df['ISTD_Conc_[nM]'], errors='coerce')
        ISTD_Annot_df[istd_conc_name] = pd.to_numeric(ISTD_Annot_df[istd_conc_name], errors='coerce')

        #Remove whitespace for each string column
        ISTD_Annot_df = MS_Template.remove_whiteSpaces(ISTD_Annot_df)

        #Close the workbook
        wb.close()
        
        return(ISTD_Annot_df)

    def __validate_ISTD_Annot_Sheet(self,worksheet):
        #Check if the sheet has been tampled
        if worksheet["A2"].value != "Transition_Name_ISTD":
            if self.__logger:
                self.__logger.error('The ISTD_Annot sheet is missing the column Transition_Name_ISTD at position A2.')
            if self.__ingui:
                print('The ISTD_Annot sheet is missing the column Transition_Name_ISTD at position A2.',flush=True)
            sys.exit(-1)

        if worksheet["E3"].value != "ISTD_Conc_[nM]":
            if self.__logger:
                self.__logger.error('The ISTD_Annot sheet is missing the column ISTD_Conc_[nM] at position E3.')
            if self.__ingui:
                print('The ISTD_Annot sheet is missing the column ISTD_Conc_[nM] at position E3.',flush=True)
            sys.exit(-1)

        if worksheet["F2"].value != "Custom_Unit":
            if self.__logger:
                self.__logger.error('The ISTD_Annot sheet is missing the column Custom_Unit at position F2.')
            if self.__ingui:
                print('The ISTD_Annot sheet is missing the column Custom_Unit at position F2.',flush=True)
            sys.exit(-1)

        if worksheet["F3"].value in ["[M]","[mM]","[uM]","[nM]","[pM]",
                                     "[M] or [mmol/mL]", "[mM] or [umol/mL]",
                                     "[uM] or [nmol/mL]", "[nM] or [pmol/mL]",
                                     "[pM] or [fmol/mL]"]:
            if self.__logger:
                self.__logger.error('Sheet ISTD_Annot\'s column Custom_Unit option ' +
                                     worksheet["F3"].value + ' ' + 
                                    'is no longer accepted in MSOrganiser. ' +
                                    'Please use a later version of MSTemplate_Creator (above 1.0.3).')
            if self.__ingui:
                print('Sheet ISTD_Annot\'s column Custom_Unit option ' +
                       worksheet["F3"].value + ' ' + 
                      'is no longer accepted in MSOrganiser. ' +
                      'Please use a later version of MSTemplate_Creator (above 1.0.3).', 
                      flush=True)
            sys.exit(-1)

        if worksheet["F3"].value not in ["[M]","[mM]","[uM]","[nM]","[pM]",
                                         "[M] or [umol/uL]", "[mM] or [nmol/uL]",
                                         "[uM] or [pmol/uL]", "[nM] or [fmol/uL]",
                                         "[pM] or [amol/uL]"]:
            if self.__logger:
                self.__logger.error('Sheet ISTD_Annot\'s column Custom_Unit option ' +
                                    worksheet["F3"].value + ' is invalid.')
            if self.__ingui:
                print('Sheet ISTD_Annot\'s column Custom_Unit option ' +
                      worksheet["F3"].value + ' is invalid.', 
                      flush=True)
            sys.exit(-1)

    def Read_Sample_Annot_Sheet(self,MS_FilePathList=[]):
        """Read the excel sheet Sample_Annot as a pandas data frame

        Args:
            MS_FilePathList (list): A list of MRM transition name file names. 
           
        Note:
            The list of MRM transition name file names names is to help the program properly filter 
            the Sample annotation such that we only pick rows whose Data_File_Name values is in the list. 
            Currently, our input is set as [os.path.basename(self.MS_FilePath)] from MSAnalysis.

        Returns:
            Sample_Annot_df (pandas DataFrame): A panda data frame containing the contents of Sample_Annot
        """

        #Open the excel file
        wb = self.__readExcelWorkbook()

        #Check if the excel file has the sheet "Sample_Annot"
        self.__checkExcelWorksheet_in_Workbook("Sample_Annot",wb)
        
        #Convert worksheet to a dataframe
        worksheet = wb["Sample_Annot"]
        #Get the column names in the first row of the excel sheet
        cols = next(worksheet.values)[0:]
        Sample_Annot_df = pd.DataFrame(worksheet.values, columns=cols)

        #We remove the first row as the headers as been set up
        Sample_Annot_df = Sample_Annot_df.iloc[1:]
        #Reset the row index
        Sample_Annot_df = Sample_Annot_df.reset_index(drop=True)

        #Remove rows with all None, NA,NaN
        Sample_Annot_df = Sample_Annot_df.dropna(axis=0, how='all')

        #Validate the Sample_Annot sheet is valid 
        # (the columns are not remove in the excel sheet but can be empty)
        self.__validate_Sample_Annot_sheet("Sample_Annot",Sample_Annot_df)

        #We take the Sample Annotation data that can be found in the MS_FilePathList
        #Else we just take all of them
        if len(MS_FilePathList) > 0:
            Sample_Annot_df = Sample_Annot_df[Sample_Annot_df.Data_File_Name.isin(MS_FilePathList)]
            MS_FilePath_with_no_sample_annot = []

            # Check that the Filtered_Sample_Annot_df is not empty for each of the provided MS_FilePath
            # If yes, stop the program and inform the user to check the Sample Annot file
            for MS_FilePath in MS_FilePathList:
                Filtered_Sample_Annot_df = Sample_Annot_df[Sample_Annot_df.Data_File_Name.isin([MS_FilePath])]

                if(len(Filtered_Sample_Annot_df.index) == 0 ):
                    MS_FilePath_with_no_sample_annot.append(MS_FilePath)

            if(len(MS_FilePath_with_no_sample_annot) > 0 ):
                if self.__logger:
                    self.__logger.error('The "Data_File_Name" column in the Sample Annotation sheet does contain the input file name.\n' +
                                        "\n".join(MS_FilePath_with_no_sample_annot) + '\n' +
                                        'Please correct the Sample Annotation sheet or the input file name.')
                if self.__ingui:
                    print('The "Data_File_Name" column in the Sample Annotation sheet does contain the input file name.\n' + 
                          "\n".join(MS_FilePath_with_no_sample_annot) + '\n' +
                          'Please correct the Sample Annotation sheet or the input file name.',
                          flush = True)
                sys.exit(-1)

        #Remove whitespaces in column names
        Sample_Annot_df.columns = Sample_Annot_df.columns.str.strip()

        #Convert all number columns to numeric
        Sample_Annot_df['Sample_Amount'] = pd.to_numeric(Sample_Annot_df['Sample_Amount'], errors='coerce')
        Sample_Annot_df['ISTD_Mixture_Volume_[uL]'] = pd.to_numeric(Sample_Annot_df['ISTD_Mixture_Volume_[uL]'], errors='coerce')
        #print(Sample_Annot_df.info())

        #Remove columns with all None, NA,NaN
        Sample_Annot_df = Sample_Annot_df.dropna(axis=1, how='all')

        #Remove whitespace for each string column
        Sample_Annot_df = MS_Template.remove_whiteSpaces(Sample_Annot_df)

        #Close the workbook
        wb.close()

        return Sample_Annot_df

    def __validate_Sample_Annot_sheet(self,sheetname,Sample_Annot_df):

        # Check if "Raw_Data_File_Name" exists as a header in Sample_Annot_df
        # If yes, give an error and ask the user to use the latest version of
        # the MSTemplate_Creator
        # Check if the column name exists as a header in the df
        if "Raw_Data_File_Name" in Sample_Annot_df:
            if self.__logger:
                self.__logger.error('The ' + sheetname  + ' sheet contains the column "Raw_Data_File_Name". ' +
                                    'This column name is no longer accepted in MSOrganiser. ' + 
                                    'Please use a later version of MSTemplate_Creator (above 0.0.1) that ' +
                                    'uses "Data_File_Name" instead.')
            if self.__ingui:
                print('The ' + sheetname  + ' sheet contains the column "Raw_Data_File_Name". ' +
                      'This column name is no longer accepted in MSOrganiser. ' + 
                      'Please use a later version of MSTemplate_Creator (above 0.0.1) that ' +
                      'uses "Data_File_Name" instead.',
                      flush=True)
            sys.exit(-1)


        # Check if the column Data_File_Name exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('Data_File_Name',sheetname,Sample_Annot_df)

        # Check if the column Merge_Status exists as a header in Sample_Annot_df
        #self.__checkColumns_in_df('Merge_Status',sheetname,Sample_Annot_df)

        # Check if the column Sample_Name exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('Sample_Name',sheetname,Sample_Annot_df)

        # Check if the column Sample_Type exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('Sample_Type',sheetname,Sample_Annot_df)

        # Check if the column Sample_Amount exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('Sample_Amount',sheetname,Sample_Annot_df)

        # Check if the column Sample_Amount_Unit exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('Sample_Amount_Unit',sheetname,Sample_Annot_df)

        # Check if the column ISTD_Mixture_Volume_[uL] exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('ISTD_Mixture_Volume_[uL]',sheetname,Sample_Annot_df)

        # Check if the column Concentration_Unit exists as a header in Sample_Annot_df
        self.__checkColumns_in_df('Concentration_Unit',sheetname,Sample_Annot_df)

        # Check if the column Data_File_Name has empty entries and highlight them.
        if(len(Sample_Annot_df[["Data_File_Name","Sample_Name"]][Sample_Annot_df["Data_File_Name"].isna()]) > 0):
            if self.__logger:
                self.__logger.warning('There are sample names that are not associated with a data file name. They will not be used during analysis.\n' +
                                      Sample_Annot_df[["Data_File_Name","Sample_Name"]][Sample_Annot_df["Data_File_Name"].isna()].to_string(index=False) + '\n'
                                      'Ensure that both columns Data_File_Name and Sample_Name are filled for each sample.')
            if self.__ingui:
                print('There are sample names that are not associated with a data file name. They will not be used during analysis.\n' +
                      Sample_Annot_df[["Data_File_Name","Sample_Name"]][Sample_Annot_df["Data_File_Name"].isna()].to_string(index=False) + '\n'
                      'Ensure that both columns Data_File_Name and Sample_Name are filled for each sample.', 
                      flush = True)

        # Check if the column Sample_Name has empty entries and highlight them.
        if(len(Sample_Annot_df[["Data_File_Name","Sample_Name"]][Sample_Annot_df["Sample_Name"].isna()]) > 0):
            if self.__logger:
                self.__logger.warning('There are data file names that are not associated with a sample name. They will not be used during analysis.\n' +
                                      Sample_Annot_df[["Data_File_Name","Sample_Name"]][Sample_Annot_df["Sample_Name"].isna()].to_string(index=False) + '\n'
                      'Ensure that both columns Data_File_Name and Sample_Name are filled for each sample.')
            if self.__ingui:
                print('There are data file names that are not associated with a sample name. They will not be used during analysis.\n' +
                      Sample_Annot_df[["Data_File_Name","Sample_Name"]][Sample_Annot_df["Sample_Name"].isna()].to_string(index=False) + '\n'
                      'Ensure that both columns Data_File_Name and Sample_Name are filled for each sample.', 
                      flush = True)
