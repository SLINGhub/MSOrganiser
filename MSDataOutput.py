from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Font
from pandas import ExcelWriter
import os
import sys
import pandas as pd


class MSDataOutput:
    """
    A class to describe the general setup for Data Output.

    Args:
        output_directory (str): directory path to output the data
        input_file_path (str): file path of the input MRM transition name file. To be used for the output filename
        logger (object): logger object created by start_logger in MSOrganiser
        ingui (bool): if True, print analysis status to screen

    """
    def __init__(self, output_directory, input_file_path, result_name = "Results" ,logger=None, ingui=True):
        self.output_directory = output_directory
        self.output_filename = os.path.splitext(os.path.basename(input_file_path))[0]
        self.result_name = result_name
        self.logger = logger
        self.ingui = ingui
        self.writer = None

    def transpose_MSdata(MS_df,allow_multiple_istd=False):
        """Function to transpose data

        Args:
            MS_df (pandas DataFrame): A panda data frame
            allow_multiple_istd (bool): if True, display the Transition_Name_ISTD as well besides the Transition_Name

        Returns:
            MS_df (pandas DataFrame): A panda data frame with data transposed

        """
        #Transpose the data
        MS_df = MS_df.T

        if allow_multiple_istd:
            # The first two columns are the index
            #Assign the column names from first row
            colnames = MS_df.iloc[0,:].astype('str').str.strip()
            MS_df.columns = colnames
            MS_df.columns.name = None
            #We remove the first row because the column names are given
            MS_df = MS_df.iloc[1:]
            #Convert the two index (Transition_Name and Transition_Name_ISTD) to columns
            MS_df.reset_index(inplace=True)  
        else:
            #Now the first column is the index, we need to convert back to a first column
            MS_df.reset_index(level=0, inplace=True)
            #Assign the column names from first row
            colnames = MS_df.iloc[0,:].astype('str').str.strip()
            MS_df.columns = colnames
            MS_df.columns.name = None
            #We remove the first row because the column names are given
            MS_df = MS_df.iloc[1:]
            #Rename the first column to Transition Name
            MS_df.rename(columns={'Sample_Name':'Transition_Name'}, inplace=True)
            #Reset the index since we remove first row and convert numeric columns from object to numeric
            MS_df = MS_df.reset_index(drop=True)
            MS_df = MS_df.apply(pd.to_numeric, errors='ignore')

        return MS_df

    def start_writer(self):
        """Function to open a writer object"""
        self.writer = os.path.join(self.output_directory, self.output_filename)

    def df_to_file_preparation(output_option,df,transpose=False,
                               logger=None,ingui=False,
                               allow_multiple_istd=False):
        """Function to check and set up the settings needed before writing to a file.

        Args:
            output_option (str): the Output_Options value given to MSOrganiser.
            df (pandas DataFrame): A panda data frame to output
            transpose (bool): if True, transpose the dataframe first before writing to the Excel sheet
            logger (object): logger object created by start_logger in MSOrganiser
            ingui (bool): if True, print analysis status to screen
            allow_multiple_istd (bool): if True, display the Transition_Name_ISTD as well besides the Transition_Name

        Returns:
            (list): list containing:

                * df (pandas DataFrame): A panda data frame to output. Transposed if transpose is set to True
                * output_option (str): Updated output options if changes are required E.g S/N to S_to_N
        
        """

        # If df is empty we send a warning and skip the df
        if df.empty:
            if logger:
                logger.warning('%s has no data. Please check the input file',output_option)
            if ingui:
                print(output_option + ' has no data. Please check the input file',flush=True)
            return([df,output_option])

        # Replace '/' as excel cannot have this as sheet title or file name
        if output_option == 'S/N':
            output_option = output_option.replace('/','_to_')

        if transpose:
            df = MSDataOutput.transpose_MSdata(df,allow_multiple_istd)

        return([df,output_option])


    def df_to_file(self,output_option,df,transpose=False,
                   allow_multiple_istd=False):
        """Funtion to write a df to a csv file named {input file name}_{output_option}_{result filename}.csv .

        Args:
            output_option (str): the name of the result csv file. MSOrganiser puts it as the Output_Options value
            df (pandas DataFrame): A panda data frame to output to csv
            transpose (bool): if True, transpose the dataframe first before writing to the Excel sheet
            allow_multiple_istd (bool): if True, display the Transition_Name_ISTD as well besides the Transition_Name

        Note:
            For the output option S/N, it will be changed to S_to_N as filenames does not accept slashes. This is done in the function df_to_file_preparation
        
        """
        if self.writer is None:
            self.start_writer()

        [df,output_option] = MSDataOutput.df_to_file_preparation(output_option,df,transpose,self.logger,self.ingui,
                                                                 allow_multiple_istd)

        #If df is empty, just leave the function, warning has been sent to df_to_file_preparation
        if df.empty:
            return

        df.to_csv(self.writer + '_' + self.result_name + '_' + output_option + '.csv',sep=',',index=False)

class MSDataOutput_csv(MSDataOutput):
    """
    A class to describe the general setup for Data Output to csv.

    Args:
        output_directory (str): directory path to output the data
        input_file_path (str): file path of the input MRM transition name file. To be used for the output filename
        logger (object): logger object created by start_logger in MSOrganiser
        ingui (bool): if True, print analysis status to screen

    """
    def df_to_file(self,output_option,df,transpose=False,
                   allow_multiple_istd=False):
        """Funtion to write a df to an excel file.

        Args:
            output_option (str): the name of the sheet
            df (pandas DataFrame): A panda data frame to output to Excel
            transpose (bool): if True, transpose the dataframe first before writing to the Excel sheet
            allow_multiple_istd (bool): if True, display the Transition_Name_ISTD as well besides the Transition_Name
        
        Note:
            For the output option S/N, it will be changed to S_to_N as excel does not accept slashes. This is done in the function df_to_file_preparation
        
        """
        if self.writer is None:
            self.start_writer()

        [df,output_option] = MSDataOutput.df_to_file_preparation(output_option,df,transpose,self.logger,self.ingui,
                                                                 allow_multiple_istd)

        #If df is empty, just leave the function, warning has been sent to df_to_file_preparation
        if df.empty:
            return

        if(output_option in ['Sample_Annot', 'Transition_Name_Annot']) :
            csv_filename = os.path.join(self.output_directory,output_option + '.csv')
        else:
            if self.result_name == "":
                csv_filename = self.writer + '_' + output_option + '.csv'
            else:
                csv_filename = self.writer + '_' + self.result_name + '_' + output_option + '.csv'

        try:
            df.to_csv(csv_filename,sep=',',index=False)
        except Exception as e:
            print("Unable to write df to csv file name" + csv_filename,flush=True)
            print(e,flush=True)

class MSDataOutput_Excel(MSDataOutput):
    """
    A class to describe the general setup for Data Output to Excel.

    Args:
        output_directory (str): directory path to output the data
        input_file_path (str): file path of the input MRM transition name file. To be used for the output filename
        logger (object): logger object created by start_logger in MSOrganiser
        ingui (bool): if True, print analysis status to screen

    """
    def start_writer(self):
        """Function to open an Excel writer object using openpyxl"""

        #Set options for excel to not turn strings into formulas
        #From https://stackoverflow.com/questions/54094172/file-corruption-while-writing-using-pandas
        #options = {}
        #options['strings_to_formulas'] = False
        #options['strings_to_urls'] = False

        self.writer = os.path.join(self.output_directory, self.output_filename + '_' +  self.result_name + '.xlsx' )

        try:
            self.writer = pd.ExcelWriter(self.writer, engine = 'openpyxl', 
                                         #options = options,
                                         engine_kwargs = {'options': {'strings_to_formulas': False,
                                                                      'strings_to_urls' : False
                                                                      }
                                                         }
                                         )
            #self.writer = pd.ExcelWriter(self.writer, engine = 'xlsxwriter', options = options)
        except UserWarning as w:
            if self.logger:
                self.logger.warning(w)
            sys.exit(-1)
        except IOError as i:
            if self.ingui:
                print('Unable to save Excel file due to:',flush=True)
                print(i,flush=True)
                print('Please close the file if it is still open on your computer, ' +
                      'then try running the software again',flush=True)
            if self.logger:
                self.logger.error('Unable to save Excel file due to:')
                self.logger.error(i)
                self.logger.error('Please close the file if it is still open on your computer, ' +
                                  'then try running the software again')
            sys.exit(-1)
        except Exception as e:
            if self.ingui:
                print('Unable to save Excel file due to:',flush=True)
                print(e,flush=True)
            if self.logger:
                self.logger.error('Unable to save Excel file due to:')
                self.logger.error(e)
            sys.exit(-1)

    def end_writer(self):
        """Function to close an Excel writer object"""
        if self.writer.engine=="xlsxwriter":
            if self.writer.book.fileclosed:
                return()

        try:
            self.writer.save()
        except UserWarning as w:
            if self.logger:
                self.logger.warning(w)
            sys.exit(-1)
        except Exception as e:
            if self.ingui:
                print('Unable to save Excel file due to:',flush=True)
                print(e,flush=True)
            if self.logger:
                self.logger.error('Unable to save Excel file due to:')
                self.logger.error(e)
            sys.exit(-1)

    def __get_col_widths(dataframe,transpose=False,allow_multiple_istd=False):

        """Function to get the correct width to output the excel file nicely"""

        if allow_multiple_istd and not transpose:
            # Assuming multiindex on the columns but not the rows
            # First we find the maximum length of the index column   
            column_index_name_length_list = [len(column_index_name) for column_index_name in dataframe.columns.names]
            row_index_name_length_list = [len(str(dataframe.index.name))]
            row_index_value_length_list = [len(str(s)) for s in dataframe.index.values] 
            idx_max = max(column_index_name_length_list + 
                          row_index_name_length_list + 
                          row_index_value_length_list
                          )
            #Next we proceed to the other columns
            max_list = []
            for cols in dataframe.columns:
                #print(cols)
                column_name_length_list = [len(col) for col in cols]            
                column_value_length_list = [len(str(s)) for s in dataframe[cols].values]
                max_list.append(max(max(column_name_length_list), max(column_value_length_list)) + 1)

            # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
            return [idx_max] + max_list
        else:
            # Assuming multiindex is not present in both row and columns
            # First we find the maximum length of the index column
            row_index_name_length_list = [len(str(dataframe.index.name))]
            row_index_value_length_list = [len(str(s)) for s in dataframe.index.values]    
            idx_max = max(row_index_name_length_list + 
                          row_index_value_length_list
                          )
            #Next we proceed to the other columns
            max_list = []
            for col in dataframe.columns:
                #print(cols)
                column_name_length = len(col)
                column_value_length_list = [len(str(s)) for s in dataframe[col].values]
                max_list.append(max([column_name_length] + column_value_length_list) + 1)

            # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
            return [idx_max] + max_list
            #return [idx_max] + [max([len(str(s)) for s in dataframe[col].values] + [len(col)]) + 1 for col in dataframe.columns]

    def df_to_file(self,output_option,df,
                   transpose=False, allow_multiple_istd=False):
        """Funtion to write a df to an excel file.

        Args:
            output_option (str): the name of the sheet
            df (pandas DataFrame): A panda data frame to output to Excel
            transpose (bool): if True, transpose the dataframe first before writing to the Excel sheet
            allow_multiple_istd (bool): if True, display the Transition_Name_ISTD as well besides the Transition_Name
        
        Note:
            For the output option S/N, it will be changed to S_to_N as excel does not accept slashes. This is done in the function df_to_file_preparation
        
        """
        if self.writer is None:
            self.start_writer()

        [df,output_option] = MSDataOutput.df_to_file_preparation(output_option,df,
                                                                 transpose,self.logger,self.ingui,
                                                                 allow_multiple_istd)

        #If df is empty, just leave the function, warning has been sent to df_to_file_preparation
        if df.empty:
            return

        try:
            #It is a pity that index must be set to True for outputing df with MultiIndex on columns or rows
            if allow_multiple_istd and not transpose:
                df.to_excel(excel_writer=self.writer,sheet_name=output_option, 
                                index=True, merge_cells=True)
            else:
                df.to_excel(excel_writer=self.writer,sheet_name=output_option, index=False)
            worksheet = self.writer.sheets[output_option]
            #Change the font to Consolas and set first row to bold
            #Go to openpyxl -> writer -> theme.py and change minor font to Consolas
            #if self.writer.engine=="openpyxl":
                #for i in range(1,len(df.columns) + 1):
                    #cells = get_column_letter(i) + ":" + get_column_letter(i)
                    #for cell in worksheet[cells]:
                    #    cell.font = Font(name='Consolas')
                #Set the first row to bold
                #for cell in worksheet[1:1]:
                #     cell.font = Font(name='Consolas', bold = True)
 
            for i, width in enumerate(MSDataOutput_Excel.__get_col_widths(df,transpose,allow_multiple_istd)):
                if self.writer.engine=="openpyxl":
                    if allow_multiple_istd and not transpose:
                        #For this case the index needs to be displayed  
                        #print("i is " + get_column_letter(i+1))
                        worksheet.column_dimensions[get_column_letter(i+1)].width = width + 5
                    else:
                        #i = 0 is for the index which we do not need to display
                        if i==0:
                            continue
                        #print("i is " + get_column_letter(i))
                        worksheet.column_dimensions[get_column_letter(i)].width = width + 5

                if self.writer.engine=="xlsxwriter":
                    if i==0:
                        continue
                    column_width_name = get_column_letter(i) + ":" + get_column_letter(i)
                    worksheet.set_column(column_width_name, width + 1)
        except Exception as e:
            print("Unable to write df to excel file",flush=True)
            print(e,flush=True)